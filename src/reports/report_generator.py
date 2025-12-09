import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_payout_report(exhibitors, output_file="PayoutReport.xlsx"):
    rows = []
    grand_total = 0.0

    # Sort exhibitors by last name
    sorted_exhibitors = sorted(exhibitors, key=lambda ex: ex.last_name)

    for ex in sorted_exhibitors:
        # Header row
        rows.append([f"Exhibitor: {ex.last_name}, {ex.first_name}",
                     f"Entry#: {ex.entry_number}",
                     f"DOB: {ex.dob_str}",
                     f"Age: {ex.age if ex.age else '—'}"])

        exhibitor_total = 0.0

        # Class entries
        rows.append(["Class", "Placement", "Ribbon", "Payout"])
        for c in ex.classes:
            rows.append([c.class_name, c.placement, c.ribbon, f"${c.payout:.2f}"])
            exhibitor_total += c.payout

        # Total line
        rows.append([f"TOTAL for {ex.first_name} {ex.last_name}", "", "", f"${exhibitor_total:.2f}"])
        rows.append(["Signature: _____________________________", "", "", ""])
        rows.append([""])  # spacer

        grand_total += exhibitor_total

    # Grand total line
    rows.append([f"GRAND TOTAL PAID TO ALL EXHIBITORS", "", "", f"${grand_total:.2f}"])

    # Export to Excel
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Payout Report", index=False, header=False)

        ws = writer.book["Payout Report"]

        # Style: left align all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

        # Bold exhibitor header lines
        for row in ws.iter_rows():
            if row[0].value and str(row[0].value).startswith("Exhibitor:"):
                for cell in row:
                    cell.font = Font(bold=True)

        # Bold and shade grand total line
        for row in ws.iter_rows():
            if row[0].value and str(row[0].value).startswith("GRAND TOTAL"):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
